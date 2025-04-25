#!test   54%  automate boring stuff excel module.

import openpyxl
import os
os.chdir(r'C:\Users\egzin\Desktop')     # change the directory to the directory of the file we wanna inspect
wb = openpyxl.load_workbook('example.xlsx')       

sheet =wb["Sheet1"]     ## workbook  has sheets and each sheet has cells ##   

print(sheet.max_row)
print(type(sheet.max_column))  # note that its an integer 


print(tuple(sheet['A1':"C3"]))

for rowOfCellObjects in (sheet["A1":"C3"]):
    for cellObj in tuple(rowOfCellObjects):
        print(cellObj.coordinate,cellObj.value)
    print("end of row",type(cellObj.coordinate))


print("--------------------------")

print(list(sheet.columns)[1])

for celobj in list(sheet.columns)[1]:
    print(celobj.value)

#! excell test python part 2


wb = openpyxl.load_workbook("example.xlsx")

wb.sheetnames
sheet_test = wb.active         #shows us the active sheet(first one)

column_1 = list(sheet_test.columns)[1]      #naming is awesome
print(column_1)
for objects in column_1:
    print(objects.value)                          #getting value attribute out of column 1
row_1 = list(sheet_test.rows)[1]
print(row_1)
for objectss in row_1:                   #getting value attribute out of row 1
    print(objectss.value)